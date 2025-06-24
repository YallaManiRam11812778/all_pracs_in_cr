# tkinter_dynamic_polars.py
import polars as pl
import tkinter as tk
from functools import partial
from tkinter import messagebox, ttk, filedialog

from openpyxl.reader.excel import load_workbook

current_label = None
confirm_file = None

# def selected_sheets_load_to_polars(all_sheet_names,filename):
#     given_excel_file = pl.read_excel(filename,sheet_name=all_sheet_names,raise_if_empty=False)
#     df_tree=ttk.Treeview(mywin,columns=df_list)
#     df_tree.pack()
    
def UploadAction(event=None):
    global current_label
    filename_ask = filedialog.askopenfilename()
    quit_button.pack_forget()
    files.append(filename_ask)
    if files:
        if current_label is not None:
            current_label.destroy()
        current_label = tk.Label(root,text=f" Selected File ---- {files[-1]}")
        current_label.pack()
    
    # Ask for confirmation before proceeding
    confirm = messagebox.askyesno("Confirmation", f"Do you want to proceed with the file: {files[-1]}?")
    if confirm:  # If user clicks 'Yes'
        confirm_file is not None
        current_label.destroy()
        button.destroy()
        select_sheets_names(file_name=files[-1])        
    else:
        print("Operation cancelled.")  # If user clicks 'No'

# # for fileName in files:
def select_sheets_names(file_name):
    wb = load_workbook(file_name)
    w = wb.sheetnames
    for sheetname in w:
        all_sheet_names.append(sheetname)

    # Multi-select Listbox
    listbox = tk.Listbox(root, selectmode='multiple', width=50, height=10)
    for sheet_name in all_sheet_names:
        listbox.insert(tk.END, sheet_name)
    listbox.pack(pady=20, padx=20)

    
    # OK Button
    ok_button = tk.Button(root, text="OK")
    ok_button.pack(pady=10)

    #  quit_again destroy restored
    quit_button.pack()

    parameter_of_listbox = partial(store_selection,listbox,file_name,ok_button)

    ok_button.config(command = parameter_of_listbox)

files = []
all_sheet_names = []
selected_sheets = []  # List to store selected sheet names

root = tk.Tk()
root.title("Manual Excel Automation")
root.geometry("10000x10000")

if confirm_file is None:
    # File from file manager
    button = tk.Button(root, text='Open', command=UploadAction)
    button.pack(padx=10)

# # Label
# label = tk.Label(root, text="This is an enhanced Tkinter window")
# label.pack(pady=20)


# def save_data():
#     selected_checkbox = checkbox_var.get()
#     selected_radio = radio_var.get()
#     messagebox.showinfo("Save", f"Data has been saved!\nOptions: {', '.join(selected_sheets)}\nCheckbox: {selected_checkbox}\nRadio: {selected_radio}")

def quit_app():
    root.quit()

def store_selection(listbox,file_name,ok_button):
    global selected_sheets
    selected_sheets = [listbox.get(i) for i in listbox.curselection()]
    listbox.destroy()
    ok_button.destroy()
    selected_sheets_load_to_polars(selected_sheets,file_name)


"""
# Checkbox
checkbox_var = tk.BooleanVar()
checkbox = tk.Checkbutton(root, text="Check me", variable=checkbox_var)
checkbox.pack(pady=20)

# Radio buttons
radio_var = tk.StringVar(value="Option 1")
radio_frame = tk.Frame(root)
radio_frame.pack(pady=20)

radio1 = tk.Radiobutton(radio_frame, text="Option 1", variable=radio_var, value="Option 1")
radio2 = tk.Radiobutton(radio_frame, text="Option 2", variable=radio_var, value="Option 2")
radio1.pack(anchor=tk.W)
radio2.pack(anchor=tk.W)

# Buttons
save_button = tk.Button(root, text="Save", command=save_data)
save_button.pack(pady=20)
"""
quit_button = tk.Button(root, text="Quit", command=quit_app)
quit_button.pack(pady=300)

# Start the main loop
root.mainloop()
